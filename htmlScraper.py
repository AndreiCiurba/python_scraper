# Import required modules
from lxml import html
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
import time
import glob
print(glob.glob("C:/Users/ACiurba/Desktop/similarwebdownloads/*"))

def open_xcel():
    #setup excel
    file_name_xcel = u'C:/Users/ACiurba/Desktop/almostUseless/python_web_scraper/similarweb.xlsx'
    wb = load_workbook(file_name_xcel)  # Work Book
    ws = wb["Sheet1"]  # Work Sheet

    s=Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service = s)
    count = 2
    for file in glob.glob("C:/Users/ACiurba/Desktop/similarwebdownloads/*"):
        driver.get(file)
        driver.implicitly_wait(1000)

        name = driver.find_element(By.XPATH, '//*[@id="overview"]/div/div/div/div[6]/div/table/tbody/tr[1]/td[2]').text
        country = driver.find_element(By.XPATH, '//*[@id="overview"]/div/div/div/div[6]/div/table/tbody/tr[4]/td[2]').text
        url = driver.find_element(By.XPATH, '//*[@id="overview"]/div/div/div/div[1]/p[2]').text
        total_visits = driver.find_element(By.XPATH, '//*[@id="overview"]/div/div/div/div[5]/div/div/div[1]/p[2]').text
        bounce_rate = driver.find_element(By.XPATH, '//*[@id="overview"]/div/div/div/div[5]/div/div/div[2]/p[2]').text
        pages_visit = driver.find_element(By.XPATH, '//*[@id="overview"]/div/div/div/div[5]/div/div/div[3]/p[2]').text
        visit_duration = driver.find_element(By.XPATH, '//*[@id="overview"]/div/div/div/div[5]/div/div/div[4]/p[2]').text
        direct = driver.find_element(By.XPATH, '(//*[local-name() ="tspan"])[4]').text
        referrals = driver.find_element(By.XPATH, '(//*[local-name() ="tspan"])[5]').text
        social = driver.find_element(By.XPATH, '(//*[local-name() ="tspan"])[7]').text
        email = driver.find_element(By.XPATH, '(//*[local-name() ="tspan"])[8]').text
        display = driver.find_element(By.XPATH, '(//*[local-name() ="tspan"])[9]').text

        organic = driver.find_element(By.XPATH, '//*[@id="keywords"]/div/div/div[2]/div[1]/div/div[2]/div[1]/span[2]').text
        paid = driver.find_element(By.XPATH, '//*[@id="keywords"]/div/div/div[2]/div[1]/div/div[2]/div[2]/span[2]').text


        distributor_location = 'A' + str(count)
        ws[distributor_location] = name

        url_location = 'B' + str(count)
        ws[url_location] = url

        country_location = 'C' + str(count)
        ws[country_location] = country

        total_visits_location = 'D' + str(count)
        ws[total_visits_location] = total_visits

        visit_duration_location = 'E' + str(count)
        ws[visit_duration_location] = visit_duration

        pages_visit_location = 'F' + str(count)
        ws[pages_visit_location] = pages_visit

        bounce_rate_location = 'G' + str(count)
        ws[bounce_rate_location] = bounce_rate

        direct_location = 'H' + str(count)
        ws[direct_location] = direct

        email_location = 'I' + str(count)
        ws[email_location] = email

        referrals_location = 'J' + str(count)
        ws[referrals_location] = referrals

        social_location = 'K' + str(count)
        ws[social_location] = social

        organic_location = 'L' + str(count)
        ws[organic_location] = organic

        paid_location = 'M' + str(count)
        ws[paid_location] = paid


        wb.save(file_name_xcel)

        count = count + 1




open_xcel()
